### MetatadataExtraction_mapsize.py ### 10/02/2020

versie_nummer = '>>> VERSIE 1.1 (met omvang HTTrack) <<<'

# om met files en mappen te kunnen werken
import os

# om csv-files te kunnen schrijven 
import csv

# import openpyxl: gebeurt in def build_excel_file; openpyxl is GEEN standaard Python, moet apart geïnstalleerd worden

# Functions to check if a file exists, read from a file, write to a file

def fileExists(filePath):
    exists = os.path.exists(filePath)
    return exists

# Functions for opening a file, writing & reading a line at a time, and closing the file (Kalb - p 271)

def openFileForWriting(filePath):
    fileHandle = open(filePath, 'w')
    return fileHandle

def writeALine(fileHandle, lineToWrite):
    # add a newline character '\n' at the end and write the line
    lineToWrite = lineToWrite + '\n'
    fileHandle.write(lineToWrite)

def openFileForReading(filePath):
    if not fileExists(filePath):
        print('The file, ' + filePath + ' does not exist - cannot read it.')
        return ''
    fileHandle = open(filePath, 'r')
    return fileHandle

def readALine(fileHandle):
    theLine = fileHandle.readline()

    # test EOF: if this occurs, return FALSE
    if not theLine:
        return False
    # if line ends with newline character '\n', strip that off the end
    if theLine.endswith('\n'):
        theLine = theLine.rstrip('\n')

    return theLine

def closeFile(fileHandle):
    fileHandle.close()

############################ SCREEN INPUT ############################################
def screen_input(scr_path, scr_output_file_name):
    # parameters tonen en eventueel wijzigen

    print ('*******************     METADATA EXTRACTIE     ***************'+versie_nummer)

    print()

    print ('locatie gearchiveerde websites (default is locatie script): ' + scr_path)
    scr_path_input = scr_path

    scr_akkoord = ' '
    while scr_akkoord != 'j' and scr_akkoord != 'n':
        scr_akkoord = input('Akkoord ? (j/n): ')

    while scr_akkoord == 'n':
        scr_path_input = input('geef volledig pad op voor locatie gearchiveerde websites: ')
    
        # test of scr_path_input een bestaande map is
        if os.path.isdir(scr_path_input):
            scr_akkoord = input('Akkoord ? (j/n): ')
        else:
            print('onbestaand of ongeldig pad')
            scr_akkoord = 'n'
       
    ### test of backslash of forward slash gebruikt wordt in systeem (Windows of Mac of Unix of ...)
            
    b_slashfound = scr_path.find(b_slash)   # test op  backward slash \
    if b_slashfound == -1:                  # als niet gevonden (b_slashfound = -1): forward slash (f_slash) in bf_slash zetten
        bf_slash = f_slash
    else:                                   # anders: backward slash (b_slash) in bf_slash zetten
        bf_slash = b_slash
        
    ### als bf_slash = f_slash,  alle eventuele b_slash veranderen in f_slash en omgekeerd
        
    if bf_slash == f_slash:                   
        path_input_replace = scr_path_input.replace(b_slash, f_slash)
    else:
        path_input_replace = scr_path_input.replace(f_slash, b_slash)
         
    # nu scr_path_input vervangen door path_input_replace, zodat de juiste slash-tekens in scr_path_input staan
    
    scr_path_input = path_input_replace
    
    print()

    scr_file_output = scr_output_file_name

    print ('naam output csv-bestand: ' + scr_file_output)

    scr_akkoord = ' '
    while scr_akkoord != 'j' and scr_akkoord != 'n':
        scr_akkoord = input('Akkoord ? (j/n): ')

    while scr_akkoord == 'n':
        scr_file_output = input('geef gewijzigde naam op voor output csv-bestand: ')
        scr_akkoord = input('Akkoord ? (j/n): ')
        
    # testen of output csv-bestand al bestaat
    
    scr_testen_of_bestaat = 'j'
    while scr_testen_of_bestaat == 'j':
        
        # bestandsnaam vormen door er al dan niet .csv achter te zetten
        
        if scr_file_output[-4:] == TEST_TEXT_TXT or scr_file_output[-4:] == TEST_TEXT_CSV:  # als de gebruiker .txt of .csv achter de bestandsnaam
                                                                                            # zou geplaatst hebben, dit eerst verwijderen
            scr_file_output_csv = scr_file_output
        else:
            scr_file_output_csv = scr_file_output+TEST_TEXT_CSV

        scr_path_output = scr_path_input+bf_slash+scr_file_output_csv
        
        if os.path.isfile(scr_path_output):
            
            # als output bestand al bestaat: vragen of het mag overschreven worden j of n
            
            scr_overschrijven = input('csv-bestand bestaat al. Overschrijven? (j/n): ')
            if scr_overschrijven == 'j':
                
                # indien ja, dan doen we verder, m.a.w; we zetten scr_testen_of_bestaat op 'n'
                scr_testen_of_bestaat = 'n'
            else:
                
                # indien  nee: gewijzigde naam vragen en opnieuw testen of al bestaat 'j' of 'n' (scr_testen_of_bestaat blijft op 'j' staan)
                scr_akkoord = 'n'
                while scr_akkoord == 'n':
                    scr_file_output = input('geef gewijzigde naam op voor output csv-bestand: ')
                    scr_akkoord = input('Akkoord ? (j/n): ')
        else:
            # als output bestand nog niet bestaat: geen verdere testen, nodig, het zal door het programma gecrëeerd worden
            scr_testen_of_bestaat = 'n'
    

    print()

    print('Website gearchiveerd met: ')
    scr_scan_type = '0'
    while scr_scan_type != '1' and scr_scan_type != '2' and scr_scan_type != '3':
        scr_scan_type = input ('(1) Httrack   (2) Wget      (3) Heritrix      (1/2/3): ')

    print()

    if scr_scan_type == '1':
        print('Website gearchiveerd met Httrack')
    if scr_scan_type == '2':
        print('Website gearchiveerd met Wget')
    if scr_scan_type == '3':
        print('Website gearchiveerd met Heritrix')
        
    scr_klaar = 'n'
    while scr_klaar != '' and scr_klaar != 'S':
        scr_klaar = input('Druk ENTER om programma te starten. (S) = Stop programma: ')

    print ()
    print ('-------------------------------------------------------------------------')
    print () 
        
    return scr_klaar, scr_path_input, scr_path_output, scr_file_output, scr_file_output_csv, scr_scan_type, bf_slash

    
############################### EINDE SCREEN INPUT #########################################################

############################### SUBMAPPEN van een MAP opzoeken #############################################
def create_list_maps (dir_path_input):
    list_dir = os.listdir(dir_path_input)

    # de mappen in list-dir overzetten naar list_maps1
    list_maps = []
    nmbr_entries = len(list_dir)
    for i in range (0, nmbr_entries):
        string_list_dir = list_dir[i]
        if os.path.isdir(dir_path_input+bf_slash+string_list_dir):
            list_maps.append(string_list_dir)

    print ('gevonden mappen = ', list_maps)
    print ()

    # aantal gevonden mappen        
    nmbr_maps = len(list_maps)
    return list_maps, nmbr_maps

############################### EINDE SUBMAPPEN van een MAP opzoeken #####################################

############################### FILES in een MAP opzoeken #####################################
def create_list_files (dir_path_input):
    list_dir = os.listdir(dir_path_input)

    # de files in list-dir overzetten naar list_files
    list_files = []
    nmbr_entries = len(list_dir)
    for i in range (0, nmbr_entries):
        string_list_dir = list_dir[i]
        if os.path.isfile(dir_path_input+bf_slash+string_list_dir):
            list_files.append(string_list_dir)

    print ('gevonden bestanden = ', list_files)
    print ()

    # aantal gevonden files        
    nmbr_files = len(list_files)
    return list_files, nmbr_files

############### .warc file zoeken en grootte (warcsize) bepalen (als niet gevonden staat grootte op 0)#####

def getsize_warcfile (dir_path_input, list_files, nmbr_files):
    #print('dir_path_input = ',dir_path_input)
    warcsize = 0
    warctext = ''
    for j in range (0, nmbr_files):
        scan2_3_text_warc_found = list_files[j].find(scan2_3_text_warc)
        if scan2_3_text_warc_found != -1:
            warcsize = os.path.getsize(dir_path_input+bf_slash+list_files[j])
            #print('dir_path_input + / + list_files[j] = ',dir_path_input+bf_slash+list_files[j])
            j = nmbr_files
    if warcsize > 0:
        warctext = ' bytes'
        if warcsize > 1024:
            warcsize = warcsize/1024
            warcsize = round(warcsize, 2)
            warctext = ' kB'
            if warcsize > 1024:
                warcsize = warcsize/1024
                warcsize = round (warcsize, 1)
                warctext = ' MB'
    return warcsize, warctext

####################### EINDE warc file grootte bepalen ############################################


############################## get_map_size ###############################################"""
def get_map_size(path_input):
    
    list_maps = []
    list_files = []
    list_path_maps = []
    list_path_files = []
    pointer_maps = -1
    nmbr_maps = 0

### 1e keer: scan_path_input = path_input (map 0)

    scan_path_input = path_input
   
### volgende keren: scan_path_input+'/'+map0+'/' + ... +'/'+mapi    
    while pointer_maps < nmbr_maps:
            
        
#------------------------ SUBMAPPEN van een MAP opzoeken --------------------------------
    
        list_dir = os.listdir(scan_path_input)

        # de mappen in list-dir overzetten naar list_maps en de files overzetten naar list_files
        # tegelijk twee parallelle tabellen opvullen met pathname voor de mappen in list_maps, en voor de files in list_files

        nmbr_entries = len(list_dir)
        for i in range (0, nmbr_entries):
            string_list_dir = list_dir[i]
            if os.path.isdir(scan_path_input+bf_slash+string_list_dir):
                list_maps.append(string_list_dir)
                list_path_maps.append(scan_path_input)
            else:
                if os.path.isfile(scan_path_input+bf_slash+string_list_dir):
                    list_files.append(string_list_dir)
                    list_path_files.append(scan_path_input)

#-------------------------- EINDE SUBMAPPEN van een MAP opzoeken -----------------------
                    
        nmbr_maps = len(list_maps)
        
        pointer_maps = pointer_maps+1
      
        if pointer_maps < nmbr_maps:
            scan_path_input = list_path_maps[pointer_maps] + bf_slash + list_maps[pointer_maps]
       

### nu listfiles doorlopen, telkens correcte pad opbouwen en dan getsize
### telkens samentellen bij totaal
### eindtotaal = size map0 (oorspronkelijke map in path_input)

    map_size = 0
    nmbr_files = len(list_files)

    for j in range (0, nmbr_files):
        file_size = os.path.getsize(list_path_files[j]+bf_slash+list_files[j])
        map_size = map_size + file_size

    return map_size
###################### einde get_map_size ##########################################

############################### SCAN TYPE 1 (Httrack) ####################################################
def scan_type1_routine (list_maps1, nmbr_maps1):
    ###  doorlopen van de lijst met gevonden mappen

    for i in range (0, nmbr_maps1):
        
        # samenstellen van het pad met de gezochte file-naam ('scan1_file_name') in de map 'list_maps1[i]'
        
        file_path_r = path_input+bf_slash+list_maps1[i]+bf_slash+scan1_file_name
        
        # map_path_r gaan we nodig hebben om er de omvang van te bepalen (zie verder: map_omvang = get_map_size(map_path_r)
        map_path_r = path_input+bf_slash+list_maps1[i]
            
        # test of de gezochte file ('scan1_file_name) bestaat in de map
        # zo ja (fileExists = True), openen en lezen van de file en zoeken naar 'scan1_text' ( = 'launched on')
        if fileExists (file_path_r):
            file_handle_r = openFileForReading (file_path_r)
            the_line = True
            scan1_text_found = -1
            
            # hieronder lezen we de gezochte file 'scan1_file_name' zolang 'the_line' NIET = False;
            # (bij end of file wordt er 'False' gezet in 'the_line' door de functie 'readALine')
            while the_line != False:
                               
                # file_path_r al in txt bestand schrijven, vooraleer een read te doen; als fout op read weten we waar die zich voordeed!!
                # door géén newline character toe te voegen moet verderop de regel gewoon aangevuld worden
                file_handle_w.write(file_path_r)
                
                the_line = readALine (file_handle_r)
                                     
                # scan de ingelezen regel (= 'the_line) om te zoeken naar 'scan1_text'
                # resultaat van de scan is de beginpositie (staat in 'scan1_text_found') in de ingelezen regel van de gezochte 'scan1_text'
                
                scan1_text_found = the_line.find(scan1_text)
                
                # als de 'scan1_text' NIET voorkomt in de regel, geeft de Python-functie '.find(scan1_text) de waarde '-1' terug
                # als 'scan1_text_found' VERSCHILLEND is van -1, dan hebben we de 'scan1_text' gevonden en scannen we the_line om Software, Datum en URL te vinden
                # in dat geval moeten we niet verder lezen, en plaatsen we 'the_line' op False om uit de 'while' te springen
                # we vullen de output-zones in en schrijven een regel in de output-file
                
                # VOORBEELD van the_line: "HTTrack3.23+swf launched on Tue, 08 Jul 2003 10:27:11 at http://users.pandora.be/andre.moreau +*.css ..."
                
                # Opmerking: posities in een string beginnen vanaf 0; dus 'HTTrack3.23' begint op positie 0 (software_from = 0)
                # als eindpositie moet er 1 positie meer gegeven worden; dus sofware_to is de positie waar de '+' staat na HTTrack.23
                # met die begin- en eindpositie wordt dan een 'slice' uit de string the_line genomen (zie verderop: row_url = the_line[url_from:url_to] )
                # analoog voor de andere velden
                
                if scan1_text_found != -1:
                    software_from = 0
                    software_to = the_line.find('+')
                    
                    # zoeken op komma+spatie (', '); the_line.find(', ') geeft de eerste positie waar ', ' voorkomt (m.a.w de positie waar de komma staat)
                    # daar 2 bij optellen (+2) om de beginpositie van de datum te vinden (de 0 van 08 Jul 2003 in het voorbeeld). Analoog voor de andere velden
                    date_from =  the_line.find(', ')+2  
                    date_to = date_from + 11
                    
                    url_from = the_line.find('at ')+3
                    url_to = the_line.find(' +')

                    row_pad = file_path_r
                    row_url = the_line[url_from:url_to]             # 'slice' nemen uit the_line om url te vinden, beginnend op url_from, tot url_to (= 1 pos meer)
                    row_date = the_line[date_from:date_to]
                    row_software = the_line[software_from:software_to]

                    # schrijf een regel in de .txt file; het veld row_pad (file_path_r) is al ingevuld geworden in het txt bestand; hier wordt de regel verder aangevuld
                    writeALine(file_handle_w,', '+row_url+', '+row_date+', '+row_software)
                    
                    # vul een tabel in en schrijf een regel in de .csv file
                    # voor Httrack : veld 'omvang' niet invullen (maar wel blanco zetten wegens '<OMVANG>' ingevuld bij defaults 
                    #row_csv_velden [row_index_omvang] = '' ######## TEST MAP-SIZE HIER INGEVOEGD
                    
                    map_omvang = get_map_size(map_path_r)
                    
                    if map_omvang > 0:
                        omvangtext = ' bytes'
                        if map_omvang > 1024:
                            map_omvang = map_omvang/1024
                            map_omvang = round(map_omvang, 2)
                            omvangtext = ' kB'
                            if map_omvang > 1024:
                                map_omvang = map_omvang/1024
                                map_omvang = round (map_omvang, 1)
                                omvangtext = ' MB'
                                if map_omvang > 1024:
                                    map_omvang = map_omvang/1024
                                    map_omvang = round (map_omvang, 1)
                                    omvangtext = ' GB'
                    if map_omvang != 0:
                        row_csv_velden [row_index_omvang] = str(map_omvang)+omvangtext
                    else:
                        row_csv_velden [row_index_omvang] = ''
                    
                    row_csv_velden [row_index_pad] = row_pad
                    row_csv_velden [row_index_snapshot] = row_text_snapshot + row_url + row_text_genomen_op + row_date
                    row_csv_velden [row_index_datering] = row_date
                    row_csv_velden [row_index_url] = row_url
                    row_csv_velden [row_index_geharvest] = row_text_geharvest + row_software
                    row_csv_velden [row_index_sitestructuur] = row_text_sitestructuur + row_text_indexhtml  #'row_text_indexhtml' als httrack, 'row_text_WARC' als wget of heritrix

                    csv_outputWriter.writerow(row_csv_velden)
                                        
                    the_line = False
                    
            # sluit de gelezen file        
            closeFile (file_handle_r)
############################# EINDE SCAN TYPE 1 (Httrack) #######################################################

############################# SCAN TYPE 2 (Wget) ######################################################

def scan_type2_routine (list_maps1, nmbr_maps1):
    
    ###  doorlopen van de lijst met gevonden mappen

    for i in range (0, nmbr_maps1):
        row_software = ''
        row_date = ''
        row_url = ''
        
        # samenstellen van het pad met de gezochte file-naam ('scan2_file_name' (= wget_logfile.txt)) in de map 'list_maps1[i]+bf_slash+scan2_submap_metadata'
        # (scan2_submap_metadata = 'metadata', scan2_file_name = 'wget_logfile.txt')
        
        file_path_r = path_input+bf_slash+list_maps1[i]+bf_slash+scan2_submap_metadata+bf_slash+scan2_file_name
        
        # ondertusen ook path bepalen voor scan2_submap_data; straks nodig om omvang van .warc file te bepalen (zie verder)
        map_path_data = path_input+bf_slash+list_maps1[i]+bf_slash+scan2_submap_data
            
        # test of de gezochte file ('scan2_file_name) bestaat in de submap
        # zo ja (fileExists = True), openen en lezen van de file en zoeken naar 'scan2_text_software/date/url'
        if fileExists (file_path_r):
            file_handle_r = openFileForReading (file_path_r)
           
            scan2_text_software_found = -1
            scan2_text_date_found = -1
            scan2_text_url_found = -1
            software_from = 0
            software_to = 0
            date_from = 0
            date_to = 0
            url_from = 0
            url_to = 0

            the_line = True
            
            # hieronder lezen we de gezochte file 'scan2_file_name' zolang 'the_line' NIET = False;
            # (bij end of file wordt er 'False' gezet in 'the_line' door de functie 'readALine')
            while the_line != False:
                the_line = readALine (file_handle_r)
                
                # scan de ingelezen regel (= 'the_line) om te zoeken naar 'scan2_text_sofware, url en starttime
                # resultaat van de scan is de beginpositie (staat in 'scan2_text_xyz_found') in de ingelezen regel van de gezochte 'scan2_text_xyz'
                
                scan2_text_software_found = the_line.find(scan2_text_software)
                scan2_text_date_found = the_line.find(scan2_text_date)
                scan2_text_url_found = the_line.find(scan2_text_url)
                
                # als de 'scan2_text_...' NIET voorkomt in de regel, geeft de Python-functie '.find(scan1_text) de waarde '-1' terug
                # als 'scan1_text_found' VERSCHILLEND is van -1, dan hebben we de 'scan1_text' gevonden en scannen we the_line om Software of Datum of URL te vinden
                # nadat we ze alledrie gevo,nden hebben moeten we niet verder lezen, en plaatsen we 'the_line' op False om uit de 'while' te springen
                
                
                if scan2_text_software_found != -1:
                    software_from = scan2_text_software_found + 10
                    software_to = the_line.find('gecompileerd')-1
                    row_software = the_line[software_from:software_to]
                if scan2_text_date_found != -1:
                    date_from =  scan2_text_date_found + 12
                    date_to = date_from + 10
                    row_date = the_line[date_from:date_to]
                if scan2_text_url_found != -1:
                    url_from = scan2_text_url_found + 5
                    row_url = the_line[url_from:]
                ##    url_to = the_line.find(' ') -> geen url to: einde van de inputregel

                if  software_from != 0 and date_from != 0 and url_from != 0:              
                    the_line = False

            row_pad = file_path_r
            
            # we vullen de output-zones in en schrijven een regel in de output-file        
            writeALine(file_handle_w, row_pad + ', '+row_url+', '+row_date+', '+row_software)

            # vul een tabel in en schrijf een regel in de .csv file
            # eerst: zoek grootte warc-file #
            
            list_files, nmbr_files = create_list_files (map_path_data)
            warcsize, warctext = getsize_warcfile (map_path_data, list_files, nmbr_files)
                        
            if warcsize != 0:
                row_csv_velden [row_index_omvang] = str(warcsize)+warctext
            else:
                row_csv_velden [row_index_omvang] = ''
            
            row_csv_velden [row_index_pad] = row_pad
            row_csv_velden [row_index_snapshot] = row_text_snapshot + row_url + row_text_genomen_op + row_date
            row_csv_velden [row_index_datering] = row_date 
            row_csv_velden [row_index_url] = row_url
            row_csv_velden [row_index_geharvest] = row_text_geharvest + row_software
            row_csv_velden [row_index_sitestructuur] = row_text_sitestructuur + row_text_warc   #'row_text_indexhtml' als httrack, 'row_text_WARC' als wget of heritrix
            csv_outputWriter.writerow(row_csv_velden)
            
            # sluit de gelezen file        
            closeFile (file_handle_r)
    
############################# EINDE SCAN TYPE 2 (Wget) ##############################################################

############################# SCAN TYPE 3 (Heritrix) ##################################################

def scan_type3_routine (list_maps1, nmbr_maps1):
    ###  doorlopen van de lijst met gevonden mappen

    for i in range (0, nmbr_maps1):
        
        # samenstellen pad met map[i] uit list_maps1 en dan listdir -> list_maps2
        
        map1_path_r = path_input+bf_slash+list_maps1[i]

        # list_maps2 opstellen en scannen op numeric map-naam
        # scan3_submap op 'NOTFOUND' stellen voor het geval we geen numeric map vinden, en om te vermijden dat het programma verderop crasht
        scan3_submap = 'NOTFOUND'
        
        list_maps2, nmbr_maps2 = create_list_maps(map1_path_r)
        for j in range (0, nmbr_maps2):
            if list_maps2 [j].isnumeric():
                scan3_submap = list_maps2 [j]
                break
            else:
                continue
        
        # samenstellen van het pad met de gezochte file-naam ('scan3_file_name' (= crawl_log)) in de map 'list_maps1[i]/scan3_submap/scan3_sub_submap
        
        file_path_r = path_input+bf_slash+list_maps1[i]+bf_slash+scan3_submap+bf_slash+scan3_sub_submap+bf_slash+scan3_file_name

        # ondertusen ook path bepalen voor scan3_sub_submap_warcs; straks nodig om omvang van .warc file te bepalen (zie verder)
        map_path_data = path_input+bf_slash+list_maps1[i]+bf_slash+scan3_submap+bf_slash+scan3_sub_submap_warcs
         
        # test of de gezochte file ('scan3_file_name' = ) bestaat in de map
        # zo ja (fileExists = True), openen en lezen van de file en zoeken naar 'scan3_text'
        if fileExists (file_path_r):
            file_handle_r = openFileForReading (file_path_r)
            the_line = True
            scan3_text_found = -1
            
            # hieronder lezen we de gezochte file 'scan3_file_name' zolang 'the_line' NIET = False;
            # (bij end of file wordt er 'False' gezet in 'the_line' door de functie 'readALine')
            while the_line != False:
                the_line = readALine (file_handle_r)
                
                # scan de ingelezen regel (= 'the_line) om te zoeken naar 'scan3_text'
                # resultaat van de scan is de beginpositie (staat in 'scan3_text_found') in de ingelezen regel van de gezochte 'scan3_text_url1'
                
                scan3_text_found = the_line.find(scan3_text_url1)
                
                # als de 'scan3_text_url1' NIET voorkomt in de regel, geeft de Python-functie '.find(scan1_text) de waarde '-1' terug
                # als 'scan3_text_found' VERSCHILLEND is van -1, dan hebben we de 'scan3_text_url1' gevonden en scannen we the_line om  Datum en URL te vinden
                # in dat geval moeten we niet verder lezen, en plaatsen we 'the_line' op False om uit de 'while' te springen
                # we vullen de output-zones in en schrijven een regel in de output-file
                
                if scan3_text_found != -1:
                    
                    date_from = 0
                    date_to = the_line.find(scan3_text_date)
                    url_www_from = the_line.find(scan3_text_url1)+4
                    url_www_to = the_line.find(scan3_text_url2)
                    url_http_from = url_www_to+3
                    url_http_to = the_line.find('://')+3
                    
                    row_pad = file_path_r
                    row_url = the_line[url_http_from:url_http_to]+the_line[url_www_from:url_www_to]
                    row_date = the_line[date_from:date_to]
                    row_software = scan3_text_software
                    
                    #writeALine(file_handle_w, file_path_r + ', '+the_line[url_http_from:url_http_to]+the_line[url_www_from:url_www_to]+', '+the_line[date_from:date_to]+', '+scan3_text_software)

                    writeALine(file_handle_w, row_pad + ', '+row_url+', '+row_date+', '+row_software)

                    # vul een tabel in en schrijf een regel in de .csv file
                    # eerst: zoek grootte warc-file #
            
                    list_files, nmbr_files = create_list_files (map_path_data)
                    warcsize, warctext = getsize_warcfile (map_path_data, list_files, nmbr_files)
                                
                    if warcsize != 0:
                        row_csv_velden [row_index_omvang] = str(warcsize)+warctext
                    else:
                        row_csv_velden [row_index_omvang] = '' 
                    
                    row_csv_velden [row_index_pad] = row_pad
                    row_csv_velden [row_index_snapshot] = row_text_snapshot + row_url + row_text_genomen_op + row_date
                    row_csv_velden [row_index_datering] = row_date
                    row_csv_velden [row_index_url] = row_url
                    row_csv_velden [row_index_geharvest] = row_text_geharvest + row_software
                    row_csv_velden [row_index_sitestructuur] = row_text_sitestructuur + row_text_warc  ##'row_text_indexhtml' als httrack, 'row_text_WARC' als wget of heritrix

                    csv_outputWriter.writerow(row_csv_velden)
                    
                    the_line = False
                    
            # sluit de gelezen file        
            closeFile (file_handle_r)

    
############################# EINDE SCAN TYPE 3(Heritrix) ##############################################################

def build_excel_file (path_input, file_output_csv):
    print()
    print ('-------------------------------------------------------------------------')
    
    # openen csv bestand
    csvFile = open(path_input+bf_slash+file_output_csv)
    #print('exampleFile = ', csvFile)
    
    # inlezen csv bestand (geheel bestand in een keer)
    csvReader = csv.reader(csvFile)
    
       
    # opbouwen 2D-lijst met volledige inhoud csv bestand (gehele lijst in een keer)
    csvData = list(csvReader)
    
    print()
    print('csv bestand ingelezen')
    print()
    

    #print ('csvData = ',csvData)
   
    # excel module importeren
    import openpyxl
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

    print()
    print('opbouwen Excel bestand')
    print()
    
    # leeg excel bestand voorbereiden met 'Metadata' als bladnaam
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    sheet.title = sheet_name 

    # voorbereiden styles
    # dunne onderlijning
    thick = Side(border_style = "thick", color = "000000") ### '000000' = zwart, 'ffffff' = wit, 'ff0000' = rood
    thin = Side(border_style = "thin", color = "000000")
    
    # rijen en kolommen in excel bestand schrijven
    nmbr_rows = len(csvData)
    #print('nmbr_rows = ', nmbr_rows)
    
    nmbr_columns = len(row_csv_xlx)
    #print('nmbr_columns = ', nmbr_columns)
    
    for i in range (0, nmbr_rows):
        for j in range(0, nmbr_columns):
            sheet [row_csv_xlx[j][0]+str(i+1)] = csvData [i][j]
            ### hieronder: kolom C, kolom H en kolom N: teksterugloop
            if row_csv_xlx[j][0] == 'C' or row_csv_xlx[j][0] == 'H' or row_csv_xlx[j][0] == 'N':
                sheet [row_csv_xlx[j][0]+str(i+1)].alignment = Alignment(wrap_text=True)
            
    # kolombreedtes instellen en per kolom, op de eerste rij instellen: bold, underline (thin) en achtergrondkleur (groen)
    for j in range (0, nmbr_columns):
        sheet.column_dimensions[row_csv_xlx[j][0]].width = row_csv_xlx[j][1]
        sheet[row_csv_xlx[j][0]+'1'].font = Font(bold=True)
        sheet[row_csv_xlx[j][0]+'1'].border = Border(bottom=thick)
        sheet[row_csv_xlx[j][0]+'1'].border = Border(right=thin)
        sheet[row_csv_xlx[j][0]+'1'].fill = PatternFill("solid", fgColor = "669933") # groen
        
    # eerste rij blokkeren
    sheet.freeze_panes = 'A2'
    
    # excel bestand wegschrijven
    excelFileName = (path_input+bf_slash+file_output_csv[:-4]+'.xlsx')
    
    print()
    print('schrijven Excel bestand')
    print()
    
    wb.save(excelFileName)
    print ('Excel bestand klaar : ',excelFileName)
    print()

    # excel bestand sluiten
    csvFile.close()
    
    klaar = input('ENTER = stop programma ')

### START PROGRAMMA ################################################################
    
# initialisatie parameters
# current path ophalen
path = os.getcwd ()
path_input = ' '
path_output = ' '

# programma zal testen of forward slash / of backward slash \ (\\ in Python) moet gebruikt worden;
# dat zal in bf_slash geplaats worden, en overal waar er een 'pad' moet opgebouwd worden
# zal dan bf_slash gebruikt worden
f_slash = '/'
b_slash = '\\'
bf_slash = ''

# parameters
output_file_name = 'ISAD_beschrijvingen'
output_txt_name = 'Metadata_output'
sheet_name = 'bestandsdeel'

TEST_TEXT_TXT = '.txt'
TEST_TEXT_CSV = '.csv'

scan_type = ' '


scan1_file_name = 'hts-log.txt'
scan1_text = 'launched on'

scan2_file_name         = 'wget_logfile.txt'
scan2_submap_metadata   = 'metadata'
scan2_text_software     = 'Software'
scan2_text_url          = 'URL'
scan2_text_date         = 'Start time'

scan2_submap_data       = 'data'

scan2_3_text_warc = '.warc'

scan3_file_name = 'crawl.log'
scan3_sub_submap = 'logs'
scan3_text_software = 'Heritrix-3.4'
scan3_text_url1 = 'dns:'
scan3_text_url2 = ' P '
scan3_text_date = 'T'

scan3_sub_submap_warcs = 'warcs'

row_csv_titels = ['Instellingsnaam - BA',
                  'Referentie - IN',
                  'Titel - TI',
                  'Datering (vrije tekst) - DO',
                  'Niveau gv',
                  'Omvang - DA',
                  'Archiefvormer - VV',
                  'Verwerving - VM',
                  'Inschrijvingsnummer - ac',
                  'Is deel van - bt',
                  'Onderwerpstrefwoord Soort onderwerp - io',
                  'Voorwaarden voor raadpleging - F2',
                  'Voorwaarden voor reproductie - RO',
                  'Fysieke kenmerken - PD',
                  'Aantekeningen op']

row_csv_velden = ['Amsab-Instituut voor Sociale Geschiedenis',
                  '<PAD>',                                              # wordt overschreven in programma
                  'Snapshot van <URL>, genomen op <DATUM>',             # wordt overschreven in programma
                  '<DATUM>',                                            # wordt overschreven in programma
                  'bestanddeel',
                  '<OMVANG>',                                           # wordt overschreven in programma
                  '<URL>',                                              # wordt overschreven in programma
                  'Geharvest met <SOFTWARE>',                           # wordt overschreven in programma
                  '',
                  '',
                  'concept',
                  'raadpleegbaar in de leeszaal',
                  'unknown rightsholder',
                  'gearchiveerde website raadpleegbaar via <index.html>/<WARC>',  # wordt overschreven: 'index.html' als httrack, 'WARC' als wget of heretrix
                  '']
row_text_snapshot       = 'Snapshot van '
row_text_genomen_op     = ', genomen op '
row_text_geharvest      = 'Geharvest met '
row_text_sitestructuur  = 'gearchiveerde website raadpleegbaar via '
row_text_omvang         = ''                                            # ingevuld met 'bytes' of 'kB' of 'MB'

row_text_indexhtml      = 'index.html'
row_text_warc           = 'WARC'

row_index_pad           = 1
row_index_snapshot      = 2
row_index_datering      = 3
row_index_omvang        = 5
row_index_url           = 6
row_index_geharvest     = 7
row_index_sitestructuur = 13

                   
row_csv_xlx =   [['A', 8.86],
                 ['B', 8.43],
                 ['C', 38.29],
                 ['D', 8.43],
                 ['E', 16.57],
                 ['F', 8.86],
                 ['G', 15.29],
                 ['H', 28.26],
                 ['I', 8.43],
                 ['J', 8.43],
                 ['K', 8.43],
                 ['L', 25.71],
                 ['M', 20.0],
                 ['N', 37.71],
                 ['O', 21.14]]
                 

### screen input
    
klaar, path_input, path_output, file_output, file_output_csv, scan_type, bf_slash = screen_input(path, output_file_name)

if klaar != 'S':

    # open output file (txt); 'Metadata_output' + '.txt' ### niet wijzigbaar op scherm; interne default ####
    file_path_w = path_input+bf_slash+output_txt_name+TEST_TEXT_TXT
    
    file_handle_w = openFileForWriting (file_path_w)
    
    # open output file (csv); 
    
    csv_outputFile = open(path_input+bf_slash+file_output_csv, 'w', newline='')
    csv_outputWriter = csv.writer(csv_outputFile)
    
    ### schrijf titellijn row-csv-titels
    
    csv_outputWriter.writerow(row_csv_titels)
    
    # in alle gevallen (scan_type = 1 of 2 of 3): bepalen eerste niveau mappen ; in list_maps1 plaatsen
    list_maps1, nmbr_maps1 = create_list_maps (path_input)
    
    ###  doorlopen van de lijst met gevonden mappen; andere routine ngl(scan_type = 1, 2 of 3)

    if scan_type == '1':
        scan_type1_routine (list_maps1, nmbr_maps1)
    elif scan_type == '2':
        scan_type2_routine (list_maps1, nmbr_maps1)
    else:
        scan_type3_routine (list_maps1, nmbr_maps1)
        
    # sluit output file .txt
    closeFile (file_handle_w)
    
    # sluit output file .csv
    csv_outputFile.close()
    
    print()
    print ('Output .txt file klaar ('+output_txt_name+' in '+path_input+')')
    print ('Output .csv file klaar ('+file_output+' in '+path_input+')')
    print()

    klaar = ''
    while klaar != 'j' and klaar != 'n':
        klaar = input('Akkoord om automatisch Excel bestand aan te maken ? (j/n) : ')

    if klaar == 'n':
        print()
        klaar = input('Geen automatische aanmaak Excel bestand. Druk ENTER om programma te stoppen ')
    else:
        build_excel_file (path_input, file_output_csv)
    
    
    

